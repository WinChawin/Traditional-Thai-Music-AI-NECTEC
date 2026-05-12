"""
normalize_to_processed.py

Normalize data/interim/ → data/processed/

Rules (preserve parens — do NOT convert to '----'):
  1. Cells with note chars or '-' (no parens): all chars compact, no whitespace
     e.g. "- - ซลท"   →  "--ซลท"
          "ดํ ท ล ท" →  "ดํทลท"
  2. Cells with parens + notes/dashes: compact incl. parens
     e.g. "( รํ รํ มํ รํ" → "(รํรํมํรํ"
          "ดํ ท ล ท )"   → "ดํทลท)"
  3. Cells with only parens (empty parens) — standardize spaces:
     - "(   )"  / "(   ... )" (both parens) → "(    )"  (4 spaces inside)
     - "(   "   (only open paren)            → "(    "  (4 spaces after)
     - "   )"   (only close paren)           → "    )"  (4 spaces before)

Paren cells stay as parens — they are NOT converted to "----".
Interpretation ("instrument doesn't play this bar") is the next step
(handled in modeling/loader via the `has_paren` flag in notes.parquet).

Works on both Windows (G: drive) and Colab (/content/drive/MyDrive).
Run with: python scripts/normalize_to_processed.py
"""

import os
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ----------- Path resolution -----------
def find_project_root() -> Path:
    candidates = [
        Path("/content/drive/MyDrive/02_FindingInstrument"),       # Colab
        Path(r"G:/My Drive/02_FindingInstrument"),                 # Windows local
        Path.cwd(),                                                 # fallback
        Path(__file__).resolve().parent.parent,                     # script in scripts/
    ]
    for c in candidates:
        if (c / "data" / "interim").exists():
            return c
    raise FileNotFoundError(
        "Cannot locate project root with data/interim/. "
        "Tried: " + ", ".join(str(c) for c in candidates)
    )


# ----------- Normalization rules -----------
SECTION_KW = (
    "ชั้น", "ท่อน", "ลูกหมด", "เกริ่น", "กลับต้น", "เที่ยว", "เปลี่ยน",
    "ตับ", "ตวง", "นก", "จระเข้", "ขมิ้น", "ขมื้น", "ลงวา", "พระเจ้าเปิดโลก",
)
NOTE_RE = re.compile(r"[ดรมฟซลท][ฺํ]?|\-")
NOTE_OR_DASH = re.compile(r"[ดรมฟซลท\-]")


def is_section_text(text) -> bool:
    return any(k in str(text) for k in SECTION_KW)


def normalize_data_cell(value) -> str | None:
    """Normalize a data-row cell per the 4-case rules. Returns None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    has_open = "(" in s
    has_close = ")" in s
    has_note_or_dash = bool(NOTE_OR_DASH.search(s))

    if has_note_or_dash:
        # Case 1 + Case 2: cells with notes/dashes → compact (parens kept)
        return re.sub(r"\s+", "", s)

    # No notes / no dashes: must be empty-paren cell
    if has_open and has_close:
        return "(    )"   # case 3a — full empty paren
    if has_open:
        return "(    "    # case 3b — open paren only
    if has_close:
        return "    )"    # case 3c — close paren only

    # Truly empty (shouldn't happen since we stripped above)
    return re.sub(r"\s+", "", s)


def has_paren(cell: str) -> bool:
    return "(" in cell or ")" in cell


def classify_row(row_values: list) -> str:
    """Return one of: 'blank', 'section', 'data'.

    Caller marks r==1 as 'title' if applicable.
    """
    filled = [c for c in row_values if c is not None and str(c).strip() != ""]
    if not filled:
        return "blank"
    if len(filled) == 1 and is_section_text(filled[0]):
        return "section"
    return "data"


def tokenize_notes(cell: str) -> list[str]:
    """Tokenize note chars from cell (parens are stripped first, then notes/dashes parsed)."""
    bare = cell.replace("(", "").replace(")", "").strip()
    return NOTE_RE.findall(bare)


# ----------- Main -----------
def main() -> None:
    root = find_project_root()
    src = root / "data" / "interim"
    dst = root / "data" / "processed"
    print(f"Project root: {root}")
    print(f"Source:       {src}")
    print(f"Destination:  {dst}\n")

    dst.mkdir(parents=True, exist_ok=True)

    records: list[dict] = []
    n_files = 0

    for inst_dir in sorted(src.iterdir()):
        if not inst_dir.is_dir():
            continue
        out_inst_dir = dst / inst_dir.name
        out_inst_dir.mkdir(exist_ok=True)

        for src_path in sorted(inst_dir.iterdir()):
            if src_path.suffix != ".xlsx":
                continue
            n_files += 1
            wb = load_workbook(src_path)
            ws = wb.active

            title: str | None = None
            section: str | None = None
            bar_idx = 0

            for r_idx, row in enumerate(ws.iter_rows(), 1):
                row_values = [c.value for c in row]
                kind = classify_row(row_values)

                # Row 1: title — keep as-is in Excel, capture for DataFrame
                if r_idx == 1 and kind != "blank":
                    if row_values[0] is not None:
                        title = str(row_values[0]).strip()
                    continue

                if kind == "blank":
                    continue

                if kind == "section":
                    # section header in col A or col B — capture, leave Excel value as-is
                    for v in row_values:
                        if v is not None and str(v).strip() != "":
                            section = str(v).strip()
                            break
                    continue

                # kind == "data"
                for c_idx, cell in enumerate(row, 1):
                    if cell.value is None:
                        continue
                    raw_text = str(cell.value).strip()
                    if not raw_text:
                        continue
                    # tolerate stray section label sitting in col A or B
                    if c_idx <= 2 and is_section_text(raw_text):
                        section = str(cell.value).strip()
                        continue

                    new_val = normalize_data_cell(cell.value)
                    cell.value = new_val
                    if new_val is None:
                        continue

                    bar_idx += 1
                    paren_flag = has_paren(new_val)
                    tokens = tokenize_notes(new_val)
                    records.append({
                        "instrument": inst_dir.name,
                        "song_file": src_path.name,
                        "title": title,
                        "section": section,
                        "bar_index": bar_idx,
                        "row": r_idx,
                        "col": c_idx,
                        "cell": new_val,
                        "tokens": tokens,
                        "n_tokens": len(tokens),
                        "has_paren": paren_flag,
                    })

            out_path = out_inst_dir / src_path.name
            wb.save(out_path)
            print(f"  ✓ {inst_dir.name}/{src_path.name}")

    df = pd.DataFrame(records)
    parquet_path = dst / "notes.parquet"
    df.to_parquet(parquet_path, index=False)

    print(f"\n--- Summary ---")
    print(f"  files processed: {n_files}")
    print(f"  total bars:      {len(df)}")
    print(f"  bars with paren: {int(df.has_paren.sum())}")
    print(f"  parquet:         {parquet_path}")
    print(f"\n  n_tokens distribution:")
    for k, v in df.n_tokens.value_counts().sort_index().items():
        print(f"    {k}: {v}")


if __name__ == "__main__":
    main()
